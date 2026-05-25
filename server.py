"""
FastAPI Server — Agent Chat MVP
Serves static frontend and provides /api endpoints to interact with the ADK agent.
"""

import os
import uuid
import json
import re
import csv
import io
import logging
import pathlib
from contextlib import asynccontextmanager

from dotenv import load_dotenv
from fastapi import FastAPI, HTTPException, Request, UploadFile, File
from fastapi.staticfiles import StaticFiles
from fastapi.responses import FileResponse, StreamingResponse, JSONResponse
from pydantic import BaseModel

import PyPDF2
from docx import Document as DocxDocument
from openpyxl import load_workbook
from google import genai
from google.adk.agents import Agent
from google.adk.runners import Runner, RunConfig
from google.adk.agents.run_config import StreamingMode
from google.adk.sessions import DatabaseSessionService
from google.genai.types import Content, Part, Blob

import db

# --- Logging Setup ---
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger(__name__)

# --- Load environment ---
load_dotenv(os.path.join(os.path.dirname(__file__), "life_coach_agent", ".env"))

# --- ADK Agent Setup ---
from life_coach_agent.agent import root_agent, novel_jobs

APP_NAME = "life_coach_app"
USER_ID = "default_user"

# --- Persistent Session Service (survives server restarts) ---
_SESSION_DB_PATH = os.path.join(os.path.dirname(__file__), "agent_sessions.db")
session_service = DatabaseSessionService(db_url=f"sqlite+aiosqlite:///{_SESSION_DB_PATH}")
logger.info("Using DatabaseSessionService (persistent) at %s", _SESSION_DB_PATH)

runner = Runner(
    agent=root_agent,
    app_name=APP_NAME,
    session_service=session_service,
)


# --- FastAPI App ---
@asynccontextmanager
async def lifespan(app: FastAPI):
    """Initialize DB on startup."""
    await db.init_db()
    yield

app = FastAPI(title="Life Coach Agent MVP", lifespan=lifespan)


# --- Global Error Handler ---
@app.exception_handler(Exception)
async def global_exception_handler(request: Request, exc: Exception):
    """Catch unhandled exceptions and return a safe JSON error."""
    logger.error("Unhandled error on %s %s: %s", request.method, request.url.path, exc, exc_info=True)
    return JSONResponse(
        status_code=500,
        content={"detail": "เกิดข้อผิดพลาดภายในระบบ กรุณาลองใหม่อีกครั้ง"}
    )


# --- Pydantic Models ---
class ChatRequest(BaseModel):
    session_id: str
    message: str
    file_context: str = ""  # extracted text from uploaded file(s)
    image_refs: list = []   # list of saved image filenames for multimodal

class ChatResponse(BaseModel):
    session_id: str
    reply: str
    msg_type: str

class SessionCreate(BaseModel):
    title: str = "New Chat"


# --- API Endpoints ---

@app.post("/api/sessions")
async def api_create_session(body: SessionCreate):
    """Create a new chat session."""
    session_id = str(uuid.uuid4())
    session_data = await db.create_session(session_id, body.title)
    await session_service.create_session(
        app_name=APP_NAME,
        user_id=USER_ID,
        session_id=session_id,
    )
    return session_data


@app.get("/api/sessions")
async def api_list_sessions():
    """List all chat sessions."""
    return await db.list_sessions()


@app.delete("/api/sessions/all")
async def api_delete_all_sessions():
    """Delete all chat sessions."""
    await db.delete_all_sessions()
    return {"status": "cleared"}


@app.delete("/api/sessions/{session_id}")
async def api_delete_session(session_id: str):
    """Delete a chat session."""
    await db.delete_session(session_id)
    return {"status": "deleted"}


@app.get("/api/export")
async def api_export_data():
    """Export all user data as JSON."""
    sessions = await db.list_sessions()
    for s in sessions:
        s['messages'] = await db.get_session_messages(s['id'])
    profile = await db.get_user_profile()
    
    return {
        "profile": profile,
        "sessions": sessions
    }


@app.get("/api/sessions/{session_id}/messages")
async def api_get_messages(session_id: str):
    """Get all messages for a session."""
    return await db.get_session_messages(session_id)

@app.get("/api/profile")
async def api_get_profile():
    """Get user profile."""
    profile = await db.get_user_profile()
    return {"profile": profile}

@app.delete("/api/profile")
async def api_delete_profile():
    """Clear user profile."""
    await db.clear_user_profile()
    if hasattr(app, "profile_summary_cache"):
        app.profile_summary_cache = None
    return {"status": "cleared"}

@app.get("/api/profile_summary")
async def api_get_profile_summary():
    """Get an AI-generated summary of the user's profile."""
    profile = await db.get_user_profile()
    if not profile:
        return {"summary": "<p style='color: var(--text-muted); text-align: center;'>ยังไม่มีข้อมูลส่วนตัว<br>กรุณาเริ่มแชทเพื่อตอบแบบสอบถามเบื้องต้น</p>"}
    
    profile_text = "\n".join(profile)
    
    # Simple cache to avoid re-generating for the same profile
    if hasattr(app, "profile_summary_cache") and app.profile_summary_cache and app.profile_summary_cache.get("text") == profile_text:
        return {"summary": app.profile_summary_cache["summary"]}

    try:
        api_key = os.environ.get("GEMINI_API_KEY", "")
        if not api_key:
            api_key = os.environ.get("GOOGLE_API_KEY", "")
            
        if not api_key:
            return {"summary": "<p style='color: var(--error-color);'>API Key is missing.</p>"}
        
        client = genai.Client(api_key=api_key)
        prompt = f"""
จากข้อมูลแบบสอบถามผู้ใช้ต่อไปนี้:
{profile_text}

ให้เขียนวิเคราะห์สั้นๆ เป็นภาษาไทย (แบบเป็นกันเอง เหมือนโค้ชชีวิต) โดยสรุปเป็น bullet points ตามหัวข้อต่อไปนี้:
- 👤 บุคลิก/ลักษณะนิสัยเบื้องต้น
- 💼 อาชีพการงาน
- 💰 รายได้ (ถ้าไม่มีข้อมูล ให้ประเมินคร่าวๆ จากอาชีพ หรือระบุว่ายังไม่มีข้อมูลและควรสอบถามเพิ่ม)
- 🎯 เป้าหมายและความท้าทาย

จัดรูปแบบข้อความแบบ HTML โดยใช้ tag <ul> <li> และ <strong> (ไม่ต้องมี tag <html> หรือ <body> ไม่ต้องมี markdown ```html) 
เน้นการวิเคราะห์เชิงลึกที่ช่วยให้ผู้ใช้อยากเริ่มพูดคุยกับโค้ช
"""
        response = client.models.generate_content(
            model='gemini-2.5-flash',
            contents=prompt
        )
        html_summary = response.text.replace("```html", "").replace("```", "").strip()
        
        # Save to cache
        app.profile_summary_cache = {"text": profile_text, "summary": html_summary}
        
        return {"summary": html_summary}
    except Exception as e:
        return {"summary": f"<p style='color: var(--error-color);'>เกิดข้อผิดพลาดในการวิเคราะห์ข้อมูล: {str(e)}</p>"}


# --- Tool Label Mapping (Thai) ---
TOOL_LABELS = {
    "web_search": "🔍 กำลังค้นหาข้อมูลจากอินเทอร์เน็ต...",
    "read_url": "📄 กำลังอ่านเนื้อหาจากเว็บไซต์...",
    "generate_image": "🎨 กำลังสร้างภาพ...",
    "create_pdf_document": "📕 กำลังสร้างไฟล์ PDF...",
    "create_docx_document": "📄 กำลังสร้างเอกสาร Word...",
    "create_xlsx_document": "📊 กำลังสร้างไฟล์ Excel...",
    "generate_mcq": "📝 กำลังสร้างคำถาม...",
    "generate_novel": "📖 กำลังเริ่มสร้างนิยาย...",
    "edit_chapter": "✏️ กำลังแก้ไขบท...",
    "export_novel": "📦 กำลัง export นิยาย...",
    "list_novels": "📚 กำลังดูรายการนิยาย...",
    "execute_python_code": "🐍 กำลังรันโค้ด Python...",
    "transfer_to_agent": "🔄 กำลังส่งต่อให้ผู้เชี่ยวชาญ...",
}


def _classify_reply(agent_reply: str) -> tuple[str, str]:
    """Classify an agent reply into (msg_type, cleaned_reply)."""
    msg_type = "text"
    stripped = agent_reply.strip()

    # Strip markdown code fences
    clean = stripped
    if clean.startswith("```json"):
        clean = clean[7:]
    elif clean.startswith("```"):
        clean = clean[3:]
    if clean.endswith("```"):
        clean = clean[:-3]
    clean = clean.strip()

    FILE_TYPES = ("docx_download", "pdf_download", "xlsx_download", "image_generation", "confirm_action", "novel_in_progress", "novel_download")
    MCQ_TYPES = ("mcq", "survey")

    # Try direct JSON parse first (handles nested objects like confirm_action with details array)
    if clean.startswith("{"):
        try:
            parsed = json.loads(clean)
            if isinstance(parsed, dict) and "type" in parsed:
                t = parsed["type"]
                if t in MCQ_TYPES:
                    return t, clean
                elif t in FILE_TYPES:
                    return "file", clean
        except json.JSONDecodeError:
            pass

    # Fallback: regex for inline JSON embedded in text (flat objects only)
    json_match = re.search(
        r'\{[^{}]*"type"\s*:\s*"(?:' + "|".join(MCQ_TYPES + FILE_TYPES) + r')"[^{}]*\}',
        clean, re.DOTALL,
    )
    if json_match:
        try:
            parsed = json.loads(json_match.group(0))
            t = parsed.get("type", "")
            if t in MCQ_TYPES:
                return t, json_match.group(0)
            elif t in FILE_TYPES:
                return "file", json_match.group(0)
        except json.JSONDecodeError:
            pass

    return msg_type, agent_reply


def _build_user_content(body: ChatRequest, agent_input: str):
    """Build multimodal Content from request body."""
    parts = [Part(text=agent_input)]

    if body.image_refs:
        for img_ref in body.image_refs:
            img_path = os.path.join(UPLOAD_DIR, img_ref)
            if os.path.exists(img_path):
                ext = os.path.splitext(img_ref)[1].lower()
                mime_map = {
                    ".png": "image/png", ".jpg": "image/jpeg",
                    ".jpeg": "image/jpeg", ".gif": "image/gif",
                    ".webp": "image/webp",
                }
                mime = mime_map.get(ext, "image/png")
                with open(img_path, "rb") as img_f:
                    img_bytes = img_f.read()
                parts.append(Part(inline_data=Blob(mime_type=mime, data=img_bytes)))
                try:
                    os.remove(img_path)
                except OSError:
                    pass
    return Content(role="user", parts=parts)


async def _prepare_agent_input(body: ChatRequest) -> str:
    """Prepare agent input text with profile and file context."""
    session_id = body.session_id
    user_message = body.message

    if user_message.startswith("[ตอบแบบสอบถาม]"):
        answer = user_message.replace("[ตอบแบบสอบถาม]", "").strip()
        await db.append_user_profile(answer)

    from datetime import datetime
    current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    agent_input = f"[System Context: Current Local Time is {current_time}]\n{user_message}"

    user_profile = await db.get_user_profile()
    if user_profile:
        profile_text = ", ".join(user_profile)
        agent_input = f"[ข้อมูลอ้างอิงผู้ใช้ (จดจำไว้ตลอดการสนทนา): {profile_text}]\n\n{agent_input}"

    if body.file_context:
        agent_input = f"""[ผู้ใช้แนบไฟล์มาด้วย — เนื้อหาของไฟล์อยู่ด้านล่าง]

--- เนื้อหาไฟล์ที่แนบ ---
{body.file_context}
--- จบเนื้อหาไฟล์ ---

ข้อความจากผู้ใช้: {agent_input}"""

    # Save user-visible message
    display_msg = user_message
    if body.file_context:
        display_msg = f"📎 {user_message}" if user_message.strip() else "📎 ส่งไฟล์แนบ"
    await db.add_message(session_id, "user", display_msg, "text")

    # Ensure ADK session exists
    existing = await session_service.get_session(
        app_name=APP_NAME, user_id=USER_ID, session_id=session_id,
    )
    if not existing:
        await session_service.create_session(
            app_name=APP_NAME, user_id=USER_ID, session_id=session_id,
        )

    return agent_input


def _sse(data: dict) -> str:
    """Format a dict as an SSE data line."""
    return f"data: {json.dumps(data, ensure_ascii=False)}\n\n"


# --- SSE Streaming Chat Endpoint ---
@app.post("/api/chat/stream")
async def api_chat_stream(body: ChatRequest, request: Request):
    """Send a message and stream agent activity + response via SSE."""
    agent_input = await _prepare_agent_input(body)
    user_content = _build_user_content(body, agent_input)
    session_id = body.session_id

    async def event_generator():
        agent_reply = ""
        full_text_accumulated = ""
        call_counter = 0
        processed_length = 0
        is_thinking = False
        buffer = ""

        # Initial thinking step
        yield _sse({"type": "step", "text": "🤔 กำลังคิดวิเคราะห์..."})

        try:
            # 1. Phase 1 Orchestrator Planning Loop
            from life_coach_agent.orchestrator import orchestrator
            run_adk_runner = False
            tasks_list = []
            
            async for event_dict in orchestrator.execute_task_graph(
                user_intent=agent_input,
                session_id=session_id
            ):
                if await request.is_disconnected():
                    logger.info(f"Client disconnected during SSE stream for session {session_id}")
                    break
                    
                if event_dict.get("type") == "control" and event_dict.get("action") == "execute_runner":
                    run_adk_runner = True
                    tasks_list = event_dict.get("tasks", [])
                    continue
                    
                yield _sse(event_dict)

            if await request.is_disconnected():
                return

            # 2. Execution Layer (Executor Service wrapping ADK Runner)
            if run_adk_runner:
                execution_content = user_content
                # Inject the plan into the prompt for the Executor
                if tasks_list:
                    task_str = "Task Plan from Planner:\n" + "\n".join(f"- {t['id']}: {t['description']}" for t in tasks_list)
                    execution_content.parts.append(Part(text=f"\n\n[System Instructions for Execution]\n{task_str}\nPlease execute these tasks step by step and synthesize a final response."))

                async for event in runner.run_async(
                    user_id=USER_ID,
                    session_id=session_id,
                    new_message=execution_content,
                    run_config=RunConfig(streaming_mode=StreamingMode.SSE)
                ):
                    if await request.is_disconnected():
                        logger.info(f"Client disconnected during SSE stream for session {session_id}")
                        break
                    # --- Tool Call Request ---
                    if event.get_function_calls():
                        for fc in event.get_function_calls():
                            call_counter += 1
                            tool_name = fc.name
                            label = TOOL_LABELS.get(tool_name, f"⚙️ กำลังใช้เครื่องมือ {tool_name}...")
                            yield _sse({"type": "step", "text": label})

                            # Send tool code/args for inline display
                            try:
                                args_dict = dict(fc.args) if fc.args else {}
                                truncated_args = {}
                                for k, v in args_dict.items():
                                    if isinstance(v, str) and len(v) > 800:
                                        truncated_args[k] = v[:800] + f"\n... (truncated, {len(v)} chars total)"
                                    elif isinstance(v, list) and len(v) > 10:
                                        truncated_args[k] = v[:10]
                                    else:
                                        truncated_args[k] = v
                                code_display = json.dumps(
                                    {"function": tool_name, "arguments": truncated_args},
                                    ensure_ascii=False, indent=2,
                                )
                                yield _sse({
                                    "type": "tool_code",
                                    "tool": tool_name,
                                    "code": code_display,
                                    "call_id": call_counter,
                                })
                            except Exception:
                                pass

                    # --- Tool Response ---
                    elif event.get_function_responses():
                        for fr in event.get_function_responses():
                            # Extract a preview of what the tool returned
                            result_preview = ""
                            try:
                                resp = fr.response
                                if isinstance(resp, dict):
                                    result_preview = json.dumps(resp, ensure_ascii=False, indent=2)
                                else:
                                    result_preview = str(resp) if resp else ""
                                # Truncate long results
                                if len(result_preview) > 1000:
                                    result_preview = result_preview[:1000] + f"\n... (truncated, {len(result_preview)} chars)"
                            except Exception:
                                result_preview = "(ไม่สามารถแสดงผลลัพธ์ได้)"

                            tool_name = fr.name if hasattr(fr, "name") else "tool"
                            yield _sse({
                                "type": "tool_result",
                                "tool": tool_name,
                                "preview": result_preview,
                            })

                    # --- Text chunks (agent thinking and normal text) ---
                    if event.content and event.content.parts:
                        for part in event.content.parts:
                            if hasattr(part, "text") and part.text:
                                full_text_accumulated = part.text
                                new_text = full_text_accumulated[processed_length:]
                                if new_text:
                                    processed_length = len(full_text_accumulated)
                                    for char in new_text:
                                        buffer += char
                                        if not is_thinking and "<think>" in buffer:
                                            is_thinking = True
                                            buffer = buffer.replace("<think>", "")
                                        elif is_thinking and "</think>" in buffer:
                                            is_thinking = False
                                            parts = buffer.split("</think>")
                                            if parts[0]:
                                                yield _sse({"type": "thinking", "text": parts[0]})
                                            buffer = parts[1] if len(parts) > 1 else ""
                                            
                                        if is_thinking and len(buffer) > 20:
                                            chunk_to_emit = buffer[:-8]
                                            buffer = buffer[-8:]
                                            if chunk_to_emit:
                                                yield _sse({"type": "thinking", "text": chunk_to_emit})
                                        elif not is_thinking and len(buffer) > 20:
                                            chunk_to_emit = buffer[:-8]
                                            buffer = buffer[-8:]
                                            if chunk_to_emit:
                                                yield _sse({"type": "chunk", "text": chunk_to_emit})
            
            # Flush remaining buffer
            if is_thinking and buffer:
                yield _sse({"type": "thinking", "text": buffer})
            elif not is_thinking and buffer:
                yield _sse({"type": "chunk", "text": buffer})

            # Clean final reply for database
            agent_reply = re.sub(r'<think>.*?</think>', '', full_text_accumulated, flags=re.DOTALL).strip()

            # Classify and save
            msg_type, agent_reply = _classify_reply(agent_reply)
            await db.add_message(session_id, "agent", agent_reply, msg_type)

            # Auto-title on first message
            messages = await db.get_session_messages(session_id)
            user_msgs = [m for m in messages if m["role"] == "user"]
            if len(user_msgs) == 1:
                title = body.message[:40] + ("..." if len(body.message) > 40 else "")
                await db.update_session_title(session_id, title)

            # Send final response
            yield _sse({"type": "final", "text": agent_reply, "msg_type": msg_type})

        except Exception as e:
            logger.error("SSE stream error: %s", e, exc_info=True)
            yield _sse({"type": "error", "text": "เกิดข้อผิดพลาดภายในระบบ กรุณาลองใหม่อีกครั้ง"})

        yield _sse({"type": "done"})

    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# --- Original Chat Endpoint (fallback) ---
@app.post("/api/chat", response_model=ChatResponse)
async def api_chat(body: ChatRequest):
    """Send a message and get an agent response (non-streaming fallback)."""
    agent_input = await _prepare_agent_input(body)
    user_content = _build_user_content(body, agent_input)
    session_id = body.session_id

    agent_reply = ""
    async for event in runner.run_async(
        user_id=USER_ID,
        session_id=session_id,
        new_message=user_content,
    ):
        if event.is_final_response():
            if event.content and event.content.parts:
                agent_reply = event.content.parts[0].text or ""

    msg_type, agent_reply = _classify_reply(agent_reply)
    await db.add_message(session_id, "agent", agent_reply, msg_type)

    messages = await db.get_session_messages(session_id)
    user_msgs = [m for m in messages if m["role"] == "user"]
    if len(user_msgs) == 1:
        title = body.message[:40] + ("..." if len(body.message) > 40 else "")
        await db.update_session_title(session_id, title)

    return ChatResponse(
        session_id=session_id,
        reply=agent_reply,
        msg_type=msg_type,
    )


# --- File Upload & Text Extraction ---
UPLOAD_DIR = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_DIR, exist_ok=True)

MAX_FILE_SIZE = 10 * 1024 * 1024  # 10 MB


def extract_text_from_file(filepath: str, filename: str) -> str:
    """Extract text content from a file based on its extension."""
    ext = os.path.splitext(filename)[1].lower()

    try:
        if ext == ".txt":
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                return f.read()

        elif ext == ".csv":
            with open(filepath, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                lines = []
                for row in reader:
                    lines.append(" | ".join(row))
                return "\n".join(lines)

        elif ext == ".pdf":
            text_parts = []
            with open(filepath, "rb") as f:
                reader = PyPDF2.PdfReader(f)
                for page in reader.pages[:50]:  # limit to 50 pages
                    page_text = page.extract_text()
                    if page_text:
                        text_parts.append(page_text)
            return "\n\n".join(text_parts)

        elif ext == ".docx":
            doc = DocxDocument(filepath)
            paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
            return "\n".join(paragraphs)

        elif ext == ".xlsx":
            wb = load_workbook(filepath, read_only=True, data_only=True)
            sheets_text = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_lines = [f"[Sheet: {sheet_name}]"]
                for row in ws.iter_rows(values_only=True):
                    row_text = " | ".join(str(cell) if cell is not None else "" for cell in row)
                    if row_text.strip(" |"):
                        sheet_lines.append(row_text)
                sheets_text.append("\n".join(sheet_lines))
            wb.close()
            return "\n\n".join(sheets_text)

        else:
            return f"(ไม่รองรับไฟล์ประเภท {ext})"

    except Exception as e:
        return f"(เกิดข้อผิดพลาดในการอ่านไฟล์: {str(e)})"


# Image file extensions
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
DOC_EXTS = {".txt", ".pdf", ".docx", ".xlsx", ".csv"}
ALL_ALLOWED = DOC_EXTS | IMAGE_EXTS


@app.post("/api/upload")
async def upload_file(file: UploadFile = File(...)):
    """Upload a file — extract text for documents, or save reference for images."""
    ext = os.path.splitext(file.filename or "")[1].lower()
    if ext not in ALL_ALLOWED:
        raise HTTPException(
            status_code=400,
            detail=f"ไฟล์ประเภท {ext} ไม่รองรับ กรุณาใช้: {', '.join(sorted(ALL_ALLOWED))}"
        )

    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="ไฟล์ใหญ่เกิน 10 MB")

    file_id = str(uuid.uuid4())[:8]
    safe_name = f"{file_id}_{file.filename}"
    filepath = os.path.join(UPLOAD_DIR, safe_name)

    with open(filepath, "wb") as f:
        f.write(content)

    # --- Image files: keep on disk, return reference for multimodal ---
    if ext in IMAGE_EXTS:
        return {
            "filename": file.filename,
            "type": "image",
            "image_ref": safe_name,
            "size": len(content),
        }

    # --- Document files: extract text and clean up ---
    extracted = extract_text_from_file(filepath, file.filename or "file.txt")

    max_chars = 8000
    truncated = False
    if len(extracted) > max_chars:
        extracted = extracted[:max_chars]
        truncated = True

    try:
        os.remove(filepath)
    except OSError:
        pass

    return {
        "filename": file.filename,
        "type": "document",
        "extracted_text": extracted,
        "char_count": len(extracted),
        "truncated": truncated,
    }


# --- Serve Uploaded Image Preview ---
IMAGE_MIME = {
    ".png": "image/png", ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".gif": "image/gif", ".webp": "image/webp",
}


@app.get("/api/uploads/{filename}")
async def serve_upload_preview(filename: str):
    """Serve an uploaded image for preview in the chat."""
    filepath = os.path.join(UPLOAD_DIR, filename)
    # --- Path traversal protection ---
    if not pathlib.Path(filepath).resolve().is_relative_to(pathlib.Path(UPLOAD_DIR).resolve()):
        raise HTTPException(status_code=400, detail="Invalid filename")
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found")
    ext = os.path.splitext(filename)[1].lower()
    media = IMAGE_MIME.get(ext, "application/octet-stream")
    return FileResponse(path=filepath, media_type=media)

exports_dir = os.path.join(os.path.dirname(__file__), "exports")
os.makedirs(exports_dir, exist_ok=True)

# MIME type mapping for file downloads
MIME_TYPES = {
    ".docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    ".pdf": "application/pdf",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}

@app.get("/api/download/{filename}")
async def download_file(filename: str):
    """Download a generated file (docx, pdf, xlsx)."""
    filepath = os.path.join(exports_dir, filename)
    # --- Path traversal protection ---
    if not pathlib.Path(filepath).resolve().is_relative_to(pathlib.Path(exports_dir).resolve()):
        raise HTTPException(status_code=400, detail="Invalid filename")
    if not os.path.exists(filepath):
        raise HTTPException(status_code=404, detail="File not found")
    ext = os.path.splitext(filename)[1].lower()
    media_type = MIME_TYPES.get(ext, "application/octet-stream")
    return FileResponse(
        path=filepath,
        filename=filename,
        media_type=media_type,
    )


# --- Novel Status Polling ---
@app.get("/api/novel-status/{job_id}")
async def novel_status(job_id: str):
    """Poll the status of a background novel generation job."""
    job = novel_jobs.get(job_id)
    if not job:
        raise HTTPException(status_code=404, detail="Job not found")
    
    response = {
        "status": job["status"],
        "title": job.get("title", ""),
        "chapters": job.get("chapters", 0),
        "current_step": job.get("current_step", ""),
        "progress": job.get("progress", 0),
        "error": job.get("error", job.get("current_step", "") if job["status"] == "error" else ""),
    }
    
    if job["status"] == "done":
        filename = job.get("docx", "")
        pdf_filename = job.get("pdf", "")
        response["download"] = {
            "type": "docx_download",
            "file_id": str(uuid.uuid4()),
            "filename": filename,
            "title": job.get("title", ""),
            "chapters": job.get("chapters", 0),
            "message": f"สร้างนิยาย '{job.get('title', '')}' เรียบร้อยแล้ว! คลิกดาวน์โหลดได้เลยครับ 📖"
        }
        if pdf_filename:
            response["pdf_download"] = {
                "type": "pdf_download",
                "file_id": str(uuid.uuid4()),
                "filename": pdf_filename,
                "title": job.get("title", ""),
                "chapters": job.get("chapters", 0),
                "message": f"สร้างนิยาย '{job.get('title', '')}' (PDF) เรียบร้อยแล้ว!"
            }
    elif job["status"] == "error":
        response["error"] = job.get("error", "Unknown error")
    
    return response


# --- Serve Frontend ---
static_dir = os.path.join(os.path.dirname(__file__), "static")

@app.get("/")
async def serve_index():
    return FileResponse(os.path.join(static_dir, "index.html"))

app.mount("/static", StaticFiles(directory=static_dir), name="static")

images_dir = os.path.join(static_dir, "outputs")
os.makedirs(images_dir, exist_ok=True)
app.mount("/images", StaticFiles(directory=images_dir), name="images")
